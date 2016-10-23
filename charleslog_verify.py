'''
Author : Bhishan Bhandari
bbhishan@gmail.com

The program uses following modules:

openpyxl   pip install openpyxl
The below mentioned modules are default in python 2.x
glob
re
urllib2 

The openpyxl is used to read/operate/write excel worksheet

Glob is used for getting all the .trace files inside a directory. The program is compatible even with
 multiple .trace files in the working directory. It segregates each .trace files to seperate .xlsx output file

re is used for matching request-body from the .trace file and matching event attributes from the request-body 
string.

urllib2 is used to unquote characters such as %20 to space and so on.


'''

import openpyxl
import glob
import re
import urllib2

'''
Globally defines the cell background color.
default_worksheet for cases where event attribute is not present in the request-body
compiling regex pattern to match request-body
compiling regex pattern to match event attribute
Globally opens qa_workbook in readmode
'''
redFill = openpyxl.styles.PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

yellowFill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

default_worksheet = 'pageName not NARB'

re_pattern = re.compile("Request-Body:<<--EOF-[0-9]*-.\n(.*)")

re_action_pattern = re.compile("Event=(.*?)&") 

qa_excel_path = '/home/bhishan/bhishanworks/programmingblog/fiverr/charlesproxy/QA_Workbook_for_iOS.xlsx'

wb = openpyxl.load_workbook(qa_excel_path)


def verify_request_body(request_body, output_ws):
    '''
    Matches regex pattern to get the event attribute. Writes event value to the output file and reads the 
respective event sheet from the qa workbook. Each attribute from the qa worksheet is stored in a dictionary as
 key:value pair. Uses urllib2 to unquote the request-body. Splits the request-body 
by '&' . Iterates over this list to get key and value and makes a comparision with the dictionary of qa records.
In case attributes do not match, it is added in the output file. In case the request attribute is not present in the qa worksheet, it is added in the output file along with colored background. 
    '''
    qa_data = {}
    
    request_body = urllib2.unquote(request_body)
    for match in re.finditer(re_action_pattern, request_body):
        actions = match.groups()
    try:
        action = actions[0]
    except:
        action = default_worksheet
    
    output_max_row = output_ws.max_row
    try:
        worksheet = wb.get_sheet_by_name(action)
        output_ws.cell(row=output_max_row + 1, column = 1).value = action
    except KeyError:
        return

    qa_max_row = worksheet.max_row
    for i in range(1, qa_max_row):
        try:
            
            if type(worksheet.cell(row = i, column = 2).value) is long:
                qa_data[worksheet.cell(row = i, column = 1).value] = str(worksheet.cell(row = i, column = 2).value)
            elif type(worksheet.cell(row = i, column = 2).value) is str:
                if '-----' in worksheet.cell(row = i, column = 2).value:
                    qa_data[worksheet.cell(row = i, column = 1).value] = (worksheet.cell(row = i, column = 2).value).split('-----')
                else:
                    qa_data[worksheet.cell(row = i, column = 1).value] = worksheet.cell(row = i, column = 2).value
            else:
                qa_data[worksheet.cell(row = i, column = 1).value] = worksheet.cell(row = i, column = 2).value
        except:
            break

    splitted_request = request_body.split('&')

    for attribute in splitted_request:
        
        if '=' in attribute:
            key, value = attribute.split('=')
            
            try:
                qa_flag = False
                if type(qa_data[key]) == list:
                    for each_value in qa_data[key]:
                        if each_value == value:
                            qa_flag = True
                            break
                    
                else:
                    if qa_data[key] == value:
                        qa_flag = True
                if qa_flag == False:    
                    output_ws.append(["", key, value])
                    '''
                    cell_to_fill = 'B' + str(output_ws.max_row)
                    cell2_to_fill = 'C' + str(output_ws.max_row)
                    cell1 = output_ws.cell(cell_to_fill)
                    cell2 = output_ws.cell(cell2_to_fill)
                    cell1.fill = redFill
                    cell2.fill = redFill '''
                del qa_data[key]
            except KeyError:
                output_ws.append(["", key, value])
                cell_to_fill = 'B' + str(output_ws.max_row)
                cell2_to_fill = 'C' + str(output_ws.max_row)
                cell1 = output_ws.cell(cell_to_fill)
                cell2 = output_ws.cell(cell2_to_fill)
                cell1.fill = yellowFill
                cell2.fill = yellowFill
        else:
            key = attribute
            try:
                if qa_data[key] is not None:
                    output_ws.append(["", key, ""])
                    '''cell_to_fill = 'B' + str(output_ws.max_row)
                    cell2_to_fill = 'C' + str(output_ws.max_row)
                    cell1 = output_ws.cell(cell_to_fill)
                    cell2 = output_ws.cell(cell2_to_fill)
                    cell1.fill = redFill
                    cell2.fill = redFill  
                    cell1.fill = redFill
                    cell2.fill = redFill '''   
                del qa_data[key]
            except KeyError:
                output_ws.append(["", key, ""])
                cell_to_fill = 'B' + str(output_ws.max_row)
                cell2_to_fill = 'C' + str(output_ws.max_row)
                cell1 = output_ws.cell(cell_to_fill)
                cell2 = output_ws.cell(cell2_to_fill)
                cell1.fill = yellowFill
                cell2.fill = yellowFill
    del qa_data[None]
    for rem_key in qa_data:
        rem_value = qa_data[rem_key]
        output_ws.append(["", rem_key, rem_value])
        cell_to_fill = 'B' + str(output_ws.max_row)
        cell2_to_fill = 'C' + str(output_ws.max_row)
        cell1 = output_ws.cell(cell_to_fill)
        cell2 = output_ws.cell(cell2_to_fill)
        cell1.fill = redFill
        cell2.fill = redFill  
        cell1.fill = redFill
        cell2.fill = redFill    
                      


def read_trace_files(file_name):
    '''
    Reads individual .trace file and matches the request-body content. Creates an individual output .xlsx file.
 Iterates over each request-body and passes it onto verify_request_body along with a reference to the output file.
Finally saves the output file with same name as the input trace file but with .xlsx extension.
    '''
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    with open(file_name, 'rb') as f:
        log_text_str = f.read()
        for matches in re.finditer(re_pattern, log_text_str):
            request_body = matches.groups()
            verify_request_body(request_body[0], output_ws)  
    output_wb.save(file_name+".xlsx") 

def get_trace_files():
    '''
    Gets all the files with .trace extension in the current directory and passes each filename to read_trace_files
 method to parse the charlesproxy log.
    '''
    for file_name in glob.glob("*.trace"):
        read_trace_files(file_name)   


if __name__ == '__main__':
    get_trace_files()
